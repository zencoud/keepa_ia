"""
Servicio para generar documentos en diferentes formatos (PDF, CSV, TXT, Excel, JSON, Markdown)
Convierte Markdown generado por IA a cada formato
"""

import json
import csv
import logging
import re
from io import BytesIO, StringIO
from datetime import datetime
from typing import Dict, Any, List

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import markdown

logger = logging.getLogger(__name__)


class DocumentGenerator:
    """Servicio para generar documentos en múltiples formatos"""
    
    def __init__(self):
        """Inicializa el generador de documentos"""
        pass
    
    def generate_pdf(self, markdown_content: str, product_data: Dict[str, Any]) -> BytesIO:
        """
        Genera un PDF desde contenido Markdown
        
        Args:
            markdown_content: Contenido en Markdown generado por IA
            product_data: Datos del producto (para metadata)
            
        Returns:
            BytesIO con el PDF generado
        """
        # Convertir Markdown a HTML primero, luego a PDF
        html_content = markdown.markdown(markdown_content, extensions=['tables', 'fenced_code'])
        
        # Para PDF, usamos reportlab directamente parseando el Markdown
        return self._markdown_to_pdf(markdown_content, product_data)
    
    def _markdown_to_pdf(self, markdown_content: str, product_data: Dict[str, Any]) -> BytesIO:
        """
        Convierte Markdown a PDF minimalista (PASO 4: Parseo Técnico)
        
        Args:
            markdown_content: Contenido en Markdown generado por IA en PASO 3
            product_data: Datos del producto para metadata
            
        Returns:
            BytesIO con PDF generado
            
        Parseo robusto de:
        - Títulos (# ##)
        - Tablas (| columna |)
        - Listas (-, *)
        - Texto normal
        """
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, 
                                   rightMargin=72, leftMargin=72,
                                   topMargin=50, bottomMargin=40)
            
            styles = getSampleStyleSheet()
            h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#333333'), spaceAfter=12, spaceBefore=6)
            h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#555555'), spaceAfter=8, spaceBefore=8)
            h3_style = ParagraphStyle('H3', parent=styles['Heading3'], fontSize=11, textColor=colors.HexColor('#666666'), spaceAfter=6, spaceBefore=6)
            normal_style = styles['Normal']
            footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
            
            story = []
            
            # Parsear Markdown línea por línea (PASO 4)
            logger.info(f"[PASO 4] Parseando {len(markdown_content)} caracteres de Markdown a PDF")
            lines = markdown_content.split('\n')
            in_table = False
            table_rows = []
            table_count = 0
            
            for line in lines:
                line_stripped = line.strip()
                
                # Ignorar líneas de firma
                if not line_stripped or line_stripped == '---' or line_stripped.startswith('*Generado'):
                    if line_stripped.startswith('*Generado'):
                        continue
                    if not in_table:
                        story.append(Spacer(1, 4))
                    continue
                
                # Headers (detectar títulos)
                if line_stripped.startswith('# '):
                    story.append(Paragraph(line_stripped[2:], h1_style))
                elif line_stripped.startswith('## '):
                    story.append(Paragraph(line_stripped[3:], h2_style))
                elif line_stripped.startswith('### '):
                    story.append(Paragraph(line_stripped[4:], h3_style))
                
                # Tablas (parseo robusto)
                elif '|' in line_stripped:
                    # Ignorar líneas separadoras de tabla (|---|---|)
                    if not line_stripped.startswith('|---'):
                        if not in_table:
                            in_table = True
                            table_rows = []
                            logger.debug(f"[PASO 4] Iniciando tabla #{table_count + 1}")
                        
                        # Parsear celdas de la tabla
                        cells = [c.strip() for c in line_stripped.split('|') if c.strip()]
                        if cells:
                            table_rows.append(cells)
                else:
                    # Cerrar tabla si estábamos en una
                    if in_table and table_rows:
                        if len(table_rows) > 1:
                            table_count += 1
                            max_cols = max(len(row) for row in table_rows)
                            col_width = 5.5 * inch / max_cols if max_cols > 0 else 2.75*inch
                            
                            logger.info(f"[PASO 4] Tabla #{table_count} generada - {len(table_rows)} filas × {max_cols} columnas")
                            
                            # Normalizar filas (asegurar mismo número de columnas)
                            normalized_rows = []
                            for row in table_rows:
                                if len(row) < max_cols:
                                    row.extend([''] * (max_cols - len(row)))
                                elif len(row) > max_cols:
                                    row = row[:max_cols]
                                normalized_rows.append(row)
                            
                            table = Table(normalized_rows, colWidths=[col_width] * max_cols)
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 9),
                                ('FONTSIZE', (0, 1), (-1, -1), 8),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                                ('TOPPADDING', (0, 0), (-1, 0), 8),
                                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                                ('TOPPADDING', (0, 1), (-1, -1), 5),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)])
                            ]))
                            story.append(table)
                            story.append(Spacer(1, 10))
                        table_rows = []
                        in_table = False
                    
                    # Listas
                    if line_stripped.startswith('- ') or line_stripped.startswith('* '):
                        text = line_stripped[2:].strip()
                        text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
                        story.append(Paragraph(f"• {text}", normal_style))
                    # Texto normal
                    elif not line_stripped.startswith('#'):
                        text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', line_stripped)
                        text = re.sub(r'\*(.+?)\*', r'<i>\1</i>', text)
                        if text:
                            story.append(Paragraph(text, normal_style))
            
            # Cerrar tabla final si existe
            if in_table and table_rows and len(table_rows) > 1:
                max_cols = max(len(row) for row in table_rows)
                col_width = 5.5 * inch / max_cols if max_cols > 0 else 2.75*inch
                table = Table(table_rows, colWidths=[col_width] * max_cols)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                    ('TOPPADDING', (0, 1), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)])
                ]))
                story.append(table)
            
            # Footer
            story.append(Spacer(1, 20))
            footer = Paragraph("Generado por Keepa AI", footer_style)
            story.append(footer)
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
            logger.info(f"[PASO 4] ✓ PDF generado exitosamente con {table_count} tabla(s)")
            return buffer
            
        except Exception as e:
            logger.error(f"[PASO 4] Error generando PDF: {e}")
            # Retornar PDF mínimo con mensaje de error
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = [
                Paragraph("Error generando documento", styles['Heading1']),
                Paragraph(f"No se pudo generar el PDF correctamente. Error: {str(e)}", styles['Normal'])
            ]
            doc.build(story)
            buffer.seek(0)
            return buffer
    
    def generate_csv(self, content: Dict[str, Any], product_data: Dict[str, Any]) -> StringIO:
        """
        Genera un CSV con datos tabulares del producto
        
        Args:
            content: Contenido estructurado
            product_data: Datos del producto
            
        Returns:
            StringIO con el CSV generado
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Keepa AI - Análisis de Producto'])
        writer.writerow([])
        writer.writerow(['Fecha de Generación', datetime.now().strftime('%d/%m/%Y %H:%M')])
        writer.writerow([])
        
        # Información del Producto
        writer.writerow(['INFORMACIÓN DEL PRODUCTO'])
        writer.writerow(['Campo', 'Valor'])
        writer.writerow(['Título', product_data.get('title', 'N/A')])
        writer.writerow(['ASIN', product_data.get('asin', 'N/A')])
        writer.writerow(['Marca', product_data.get('brand', 'N/A')])
        writer.writerow([])
        
        # Precios
        writer.writerow(['ANÁLISIS DE PRECIOS'])
        writer.writerow(['Métrica', 'Valor'])
        writer.writerow(['Precio Actual (Nuevo)', f"${product_data.get('current_price_new', 0) / 100:.2f}"])
        
        if content.get('price_analysis'):
            writer.writerow(['Precio Promedio', content['price_analysis'].get('average_price', 'N/A')])
            writer.writerow(['Precio Mínimo', content['price_analysis'].get('min_price', 'N/A')])
            writer.writerow(['Precio Máximo', content['price_analysis'].get('max_price', 'N/A')])
            writer.writerow(['Tendencia', content['price_analysis'].get('trend', 'N/A')])
            if content['price_analysis'].get('historical_data') and content['price_analysis']['historical_data'] not in ['N/A', '']:
                writer.writerow([])
                writer.writerow(['HISTORIAL DETALLADO'])
                writer.writerow([content['price_analysis']['historical_data']])
        writer.writerow([])
        
        # Reputación
        writer.writerow(['REPUTACIÓN Y CALIDAD'])
        writer.writerow(['Métrica', 'Valor'])
        writer.writerow(['Rating', f"{product_data.get('rating', 0):.1f}"])
        writer.writerow(['Total de Reseñas', f"{product_data.get('review_count', 0):,}"])
        
        if content.get('reputation_analysis'):
            writer.writerow(['Reviews Positivas', content['reputation_analysis'].get('positive_reviews', 'N/A')])
        writer.writerow([])
        
        # Recomendación
        if content.get('recommendation'):
            writer.writerow(['RECOMENDACIÓN'])
            writer.writerow([content['recommendation']])
        writer.writerow([])
        
        # Notas adicionales
        if content.get('additional_notes') and content['additional_notes'] not in ['N/A', '']:
            writer.writerow(['INFORMACIÓN ADICIONAL'])
            writer.writerow([content['additional_notes']])
        
        output.seek(0)
        return output
    
    def generate_txt(self, content: Dict[str, Any], product_data: Dict[str, Any]) -> StringIO:
        """
        Genera un archivo de texto plano
        
        Args:
            content: Contenido estructurado
            product_data: Datos del producto
            
        Returns:
            StringIO con el TXT generado
        """
        output = StringIO()
        
        output.write("=" * 80 + "\n")
        output.write("KEEPA AI - ANÁLISIS DE PRODUCTO\n")
        output.write("=" * 80 + "\n\n")
        
        output.write(f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n")
        
        # Información del Producto
        output.write("-" * 80 + "\n")
        output.write("INFORMACIÓN DEL PRODUCTO\n")
        output.write("-" * 80 + "\n")
        output.write(f"Título: {product_data.get('title', 'N/A')}\n")
        output.write(f"ASIN: {product_data.get('asin', 'N/A')}\n")
        output.write(f"Marca: {product_data.get('brand', 'N/A')}\n\n")
        
        # Resumen Ejecutivo
        if content.get('executive_summary'):
            output.write("-" * 80 + "\n")
            output.write("RESUMEN EJECUTIVO\n")
            output.write("-" * 80 + "\n")
            output.write(content['executive_summary'] + "\n\n")
        
        # Análisis de Precios
        if content.get('price_analysis'):
            output.write("-" * 80 + "\n")
            output.write("ANÁLISIS DE PRECIOS\n")
            output.write("-" * 80 + "\n")
            output.write(f"Precio Actual (Nuevo): ${product_data.get('current_price_new', 0) / 100:.2f}\n")
            output.write(f"Precio Promedio: {content['price_analysis'].get('average_price', 'N/A')}\n")
            output.write(f"Precio Mínimo: {content['price_analysis'].get('min_price', 'N/A')}\n")
            output.write(f"Precio Máximo: {content['price_analysis'].get('max_price', 'N/A')}\n")
            output.write(f"Tendencia: {content['price_analysis'].get('trend', 'N/A')}\n")
            
            # Datos históricos si existen
            if content['price_analysis'].get('historical_data') and content['price_analysis']['historical_data'] not in ['N/A', '']:
                output.write(f"\n{content['price_analysis']['historical_data']}\n")
            output.write("\n")
        
        # Reputación
        if content.get('reputation_analysis'):
            output.write("-" * 80 + "\n")
            output.write("REPUTACIÓN Y CALIDAD\n")
            output.write("-" * 80 + "\n")
            output.write(f"Rating: {product_data.get('rating', 0):.1f} estrellas\n")
            output.write(f"Total de Reseñas: {product_data.get('review_count', 0):,}\n")
            output.write(f"Reviews Positivas: {content['reputation_analysis'].get('positive_reviews', 'N/A')}\n\n")
        
        # Recomendación
        if content.get('recommendation'):
            output.write("-" * 80 + "\n")
            output.write("RECOMENDACIÓN\n")
            output.write("-" * 80 + "\n")
            output.write(content['recommendation'] + "\n\n")
        
        # Notas adicionales
        if content.get('additional_notes') and content['additional_notes'] not in ['N/A', '']:
            output.write("-" * 80 + "\n")
            output.write("INFORMACIÓN ADICIONAL\n")
            output.write("-" * 80 + "\n")
            output.write(content['additional_notes'] + "\n\n")
        
        output.write("=" * 80 + "\n")
        output.write("Documento generado por Keepa AI\n")
        output.write("=" * 80 + "\n")
        
        output.seek(0)
        return output
    
    def generate_excel(self, content: Dict[str, Any], product_data: Dict[str, Any]) -> BytesIO:
        """
        Genera un archivo Excel con múltiples hojas
        
        Args:
            content: Contenido estructurado
            product_data: Datos del producto
            
        Returns:
            BytesIO con el Excel generado
        """
        buffer = BytesIO()
        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="4C51BF")
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        ws1['A1'] = "KEEPA AI - ANÁLISIS DE PRODUCTO"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:B1')
        
        ws1['A3'] = "Fecha de Generación:"
        ws1['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        ws1['A5'] = "INFORMACIÓN DEL PRODUCTO"
        ws1['A5'].font = Font(bold=True, size=12)
        
        row = 6
        ws1[f'A{row}'] = "Título"
        ws1[f'B{row}'] = product_data.get('title', 'N/A')
        row += 1
        ws1[f'A{row}'] = "ASIN"
        ws1[f'B{row}'] = product_data.get('asin', 'N/A')
        row += 1
        ws1[f'A{row}'] = "Marca"
        ws1[f'B{row}'] = product_data.get('brand', 'N/A')
        
        # Ajustar columnas
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 50
        
        # Hoja 2: Análisis de Precios
        ws2 = wb.create_sheet("Precios")
        ws2['A1'] = "ANÁLISIS DE PRECIOS"
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:B1')
        
        ws2['A3'] = "Métrica"
        ws2['B3'] = "Valor"
        ws2['A3'].fill = header_fill
        ws2['B3'].fill = header_fill
        ws2['A3'].font = header_font
        ws2['B3'].font = header_font
        
        row = 4
        ws2[f'A{row}'] = "Precio Actual (Nuevo)"
        ws2[f'B{row}'] = f"${product_data.get('current_price_new', 0) / 100:.2f}"
        
        if content.get('price_analysis'):
            row += 1
            ws2[f'A{row}'] = "Precio Promedio"
            ws2[f'B{row}'] = content['price_analysis'].get('average_price', 'N/A')
            row += 1
            ws2[f'A{row}'] = "Precio Mínimo"
            ws2[f'B{row}'] = content['price_analysis'].get('min_price', 'N/A')
            row += 1
            ws2[f'A{row}'] = "Precio Máximo"
            ws2[f'B{row}'] = content['price_analysis'].get('max_price', 'N/A')
            row += 1
            ws2[f'A{row}'] = "Tendencia"
            ws2[f'B{row}'] = content['price_analysis'].get('trend', 'N/A')
        
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 30
        
        # Hoja 3: Reputación
        ws3 = wb.create_sheet("Reputación")
        ws3['A1'] = "REPUTACIÓN Y CALIDAD"
        ws3['A1'].font = title_font
        ws3.merge_cells('A1:B1')
        
        ws3['A3'] = "Métrica"
        ws3['B3'] = "Valor"
        ws3['A3'].fill = header_fill
        ws3['B3'].fill = header_fill
        ws3['A3'].font = header_font
        ws3['B3'].font = header_font
        
        ws3['A4'] = "Rating"
        ws3['B4'] = f"{product_data.get('rating', 0):.1f}"
        ws3['A5'] = "Total de Reseñas"
        ws3['B5'] = f"{product_data.get('review_count', 0):,}"
        
        if content.get('reputation_analysis'):
            ws3['A6'] = "Reviews Positivas"
            ws3['B6'] = content['reputation_analysis'].get('positive_reviews', 'N/A')
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 30
        
        # Guardar
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def generate_json(self, content: Dict[str, Any], product_data: Dict[str, Any]) -> StringIO:
        """
        Genera un archivo JSON con todos los datos estructurados
        
        Args:
            content: Contenido estructurado
            product_data: Datos del producto
            
        Returns:
            StringIO con el JSON generado
        """
        output = StringIO()
        
        data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "generator": "Keepa AI",
                "version": "1.0"
            },
            "product": {
                "title": product_data.get('title'),
                "asin": product_data.get('asin'),
                "brand": product_data.get('brand'),
                "current_price_new": product_data.get('current_price_new', 0) / 100,
                "current_price_amazon": product_data.get('current_price_amazon', 0) / 100 if product_data.get('current_price_amazon') else None,
                "rating": product_data.get('rating'),
                "review_count": product_data.get('review_count'),
                "sales_rank_current": product_data.get('sales_rank_current')
            },
            "analysis": content
        }
        
        json.dump(data, output, indent=2, ensure_ascii=False)
        output.seek(0)
        
        return output
    
    def generate_markdown(self, content: Dict[str, Any], product_data: Dict[str, Any]) -> StringIO:
        """
        Genera un archivo Markdown con formato rico
        
        Args:
            content: Contenido estructurado
            product_data: Datos del producto
            
        Returns:
            StringIO con el Markdown generado
        """
        output = StringIO()
        
        output.write("# 📊 Keepa AI - Análisis de Producto\n\n")
        output.write(f"**Fecha de Generación:** {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n")
        output.write("---\n\n")
        
        # Información del Producto
        output.write("## 📦 Información del Producto\n\n")
        output.write(f"- **Título:** {product_data.get('title', 'N/A')}\n")
        output.write(f"- **ASIN:** `{product_data.get('asin', 'N/A')}`\n")
        output.write(f"- **Marca:** {product_data.get('brand', 'N/A')}\n\n")
        
        # Resumen Ejecutivo
        if content.get('executive_summary'):
            output.write("## 📝 Resumen Ejecutivo\n\n")
            output.write(content['executive_summary'] + "\n\n")
        
        # Análisis de Precios
        if content.get('price_analysis'):
            output.write("## 💵 Análisis de Precios\n\n")
            output.write("| Métrica | Valor |\n")
            output.write("|---------|-------|\n")
            output.write(f"| Precio Actual (Nuevo) | ${product_data.get('current_price_new', 0) / 100:.2f} |\n")
            output.write(f"| Precio Promedio | {content['price_analysis'].get('average_price', 'N/A')} |\n")
            output.write(f"| Precio Mínimo | {content['price_analysis'].get('min_price', 'N/A')} |\n")
            output.write(f"| Precio Máximo | {content['price_analysis'].get('max_price', 'N/A')} |\n")
            output.write(f"| Tendencia | {content['price_analysis'].get('trend', 'N/A')} |\n\n")
            
            # Datos históricos si existen
            if content['price_analysis'].get('historical_data') and content['price_analysis']['historical_data'] not in ['N/A', '']:
                output.write("### Historial Detallado\n\n")
                output.write(f"{content['price_analysis']['historical_data']}\n\n")
        
        # Reputación
        if content.get('reputation_analysis'):
            output.write("## ⭐ Reputación y Calidad\n\n")
            output.write("| Métrica | Valor |\n")
            output.write("|---------|-------|\n")
            output.write(f"| Rating | {product_data.get('rating', 0):.1f} ⭐ |\n")
            output.write(f"| Total de Reseñas | {product_data.get('review_count', 0):,} |\n")
            output.write(f"| Reviews Positivas | {content['reputation_analysis'].get('positive_reviews', 'N/A')} |\n\n")
        
        # Recomendación
        if content.get('recommendation'):
            output.write("## 💡 Recomendación\n\n")
            output.write(f"> {content['recommendation']}\n\n")
        
        # Notas adicionales
        if content.get('additional_notes') and content['additional_notes'] not in ['N/A', '']:
            output.write("## 📌 Información Adicional\n\n")
            output.write(f"{content['additional_notes']}\n\n")
        
        output.write("---\n\n")
        output.write("*Documento generado por Keepa AI*\n")
        
        output.seek(0)
        return output
    
    def generate_txt_from_markdown(self, markdown_content: str, product_data: Dict[str, Any]) -> StringIO:
        """Convierte Markdown a texto plano removiendo formato"""
        output = StringIO()
        
        # Convertir Markdown básico a texto
        lines = markdown_content.split('\n')
        for line in lines:
            # Remover formato Markdown
            text = line
            text = re.sub(r'^#+\s*', '', text)  # Headers
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)  # Negritas
            text = re.sub(r'\*(.+?)\*', r'\1', text)  # Cursivas
            text = re.sub(r'`(.+?)`', r'\1', text)  # Código
            text = re.sub(r'^\s*[-*]\s+', '• ', text)  # Listas
            text = re.sub(r'\|', ' | ', text)  # Tablas - separar columnas
            output.write(text + "\n")
        
        output.seek(0)
        return output
    
    def generate_csv_from_markdown(self, markdown_content: str, product_data: Dict[str, Any]) -> StringIO:
        """Extrae tablas del Markdown y genera CSV minimalista"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Parsear Markdown para encontrar tablas
        lines = markdown_content.split('\n')
        in_table = False
        
        for line in lines:
            line = line.strip()
            
            # Detectar tablas
            if '|' in line and not line.startswith('|---') and not line.startswith('---'):
                if not in_table:
                    in_table = True
                
                # Parsear fila
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                if cells:
                    writer.writerow(cells)
            elif in_table and not line:
                # Separador entre tablas
                in_table = False
            elif line.startswith('#'):
                # Agregar títulos como comentarios
                title = re.sub(r'^#+\s*', '', line).strip()
                if title and not title.startswith('---'):
                    writer.writerow([f"# {title}"])
        
        output.seek(0)
        return output
    
    def generate_excel_from_markdown(self, markdown_content: str, product_data: Dict[str, Any]) -> BytesIO:
        """Convierte Markdown a Excel minimalista"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        
        row = 1
        lines = markdown_content.split('\n')
        in_table = False
        first_table_row = True
        
        for line in lines:
            line = line.strip()
            if not line or line == '---' or line.startswith('*Generado'):
                continue
            
            # Títulos
            if line.startswith('#'):
                title = re.sub(r'^#+\s*', '', line).strip()
                ws[f'A{row}'] = title
                ws[f'A{row}'].font = title_font
                row += 1
            
            # Tablas
            elif '|' in line and not line.startswith('|---'):
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                if cells:
                    for col_idx, cell in enumerate(cells, 1):
                        cell_obj = ws.cell(row=row, column=col_idx, value=cell)
                        
                        # Primera fila de tabla = header
                        if first_table_row:
                            cell_obj.fill = header_fill
                            cell_obj.font = header_font
                    
                    first_table_row = False
                    row += 1
            
            # Reset al salir de tabla
            elif not line.startswith('|'):
                if not first_table_row:
                    first_table_row = True
                    row += 1  # Espacio después de tabla
        
        # Ajustar anchos
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_json_from_markdown(self, markdown_content: str, product_data: Dict[str, Any]) -> StringIO:
        """Convierte Markdown a JSON minimalista con datos tabulares"""
        output = StringIO()
        
        # Parsear tablas del Markdown
        tables = []
        lines = markdown_content.split('\n')
        current_table = []
        headers = []
        
        for line in lines:
            line = line.strip()
            if '|' in line and not line.startswith('|---') and not line.startswith('---'):
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                if cells:
                    if not headers:
                        headers = cells
                    else:
                        if len(cells) == len(headers):
                            row_dict = dict(zip(headers, cells))
                            current_table.append(row_dict)
            elif current_table:
                tables.append(current_table)
                current_table = []
                headers = []
        
        if current_table:
            tables.append(current_table)
        
        # Construir JSON
        data = {
            "generated_by": "Keepa AI",
            "generated_at": datetime.now().isoformat(),
            "data": tables[0] if len(tables) == 1 else tables
        }
        
        json.dump(data, output, indent=2, ensure_ascii=False)
        output.seek(0)
        return output

